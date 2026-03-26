# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import api, fields, models
from odoo.exceptions import UserError
from odoo.addons.l10n_ve_payment_config.utils.payment_gateway import (
    PaymentGatewayClient, PGConfig
)

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    Workbook = None


class VeCertificationWizard(models.TransientModel):
    _name = 've.certification.wizard'
    _description = 'Wizard de Certificación MegaSoft'

    gateway_config_id = fields.Many2one(
        've.payment.gateway.config',
        string='Configuración de Pasarela',
        required=True,
    )
    pos_session_id = fields.Many2one(
        'pos.session',
        string='Sesión POS',
        domain="[('state', '=', 'opened')]",
        help='Seleccione la sesión POS abierta donde se registrarán las transacciones de prueba.',
    )
    active_pos_session_display = fields.Char(
        string='Info Sesión',
        compute='_compute_active_pos_session',
    )

    # Datos del integrador
    ci_integrador = fields.Char(
        string='CI / RIF del Integrador',
        required=True,
        help='Ej: V12345678',
    )
    telefono_integrador = fields.Char(
        string='Teléfono del Integrador',
        required=True,
        help='Ej: 04121234567',
    )
    banco_p2c_id = fields.Many2one(
        've.payment.service.bank',
        string='Banco Comercio P2C',
        domain="[('service_id.service_code', '=', 'p2c')]",
    )
    banco_zelle_id = fields.Many2one(
        've.payment.service.bank',
        string='Banco Comercio Zelle',
        domain="[('service_id.service_code', '=', 'zelle')]",
    )

    # Resultados
    state = fields.Selection([
        ('draft', 'Pendiente'),
        ('running', 'Ejecutando...'),
        ('done', 'Completado'),
    ], default='draft')
    result_line_ids = fields.One2many(
        've.certification.result.line',
        'wizard_id',
        string='Resultados',
        readonly=True,
    )
    passed_count = fields.Integer(compute='_compute_counts', string='PASS')
    failed_count = fields.Integer(compute='_compute_counts', string='FAIL')
    total_count = fields.Integer(compute='_compute_counts', string='Total')
    export_file = fields.Binary(string='Archivo Excel', readonly=True)
    export_filename = fields.Char(string='Nombre Archivo')

    @api.depends('result_line_ids', 'result_line_ids.passed')
    def _compute_counts(self):
        for rec in self:
            lines = rec.result_line_ids
            rec.total_count = len(lines)
            rec.passed_count = len(lines.filtered('passed'))
            rec.failed_count = rec.total_count - rec.passed_count

    def _compute_active_pos_session(self):
        for rec in self:
            if rec.pos_session_id:
                rec.active_pos_session_display = '%s (ID: %s)' % (
                    rec.pos_session_id.name, rec.pos_session_id.id
                )
            else:
                # Intentar autodetectar
                session = self.env['pos.session'].search([
                    ('state', '=', 'opened'),
                ], limit=1)
                if session:
                    rec.active_pos_session_display = (
                        'Sesión disponible: %s — selecciónela arriba.' % session.name
                    )
                else:
                    rec.active_pos_session_display = (
                        'No hay sesión POS abierta. Abra una sesión antes de ejecutar las pruebas.'
                    )

    # ── Ejecución de pruebas ──────────────────────────────────────

    def action_run_tests(self):
        self.ensure_one()
        pos_session = self.pos_session_id
        if not pos_session:
            raise UserError(
                'Debe seleccionar una sesión del Punto de Venta. '
                'Seleccione una sesión POS abierta antes de ejecutar las pruebas de certificación.'
            )

        # Validar teléfono del banco P2C
        if self.banco_p2c_id and not self.banco_p2c_id.phone_number:
            raise UserError(
                'El banco P2C seleccionado (%s) no tiene teléfono del comercio configurado. '
                'Configure el teléfono en el servicio P2C antes de ejecutar las pruebas.'
                % self.banco_p2c_id.bank_id.name
            )

        # Validar cuenta destino transferencia
        banco_transf = self.env['ve.payment.service.bank'].search([
            ('service_id.gateway_config_id', '=', self.gateway_config_id.id),
            ('service_id.service_code', '=', 'transferencia'),
            ('active', '=', True),
        ], limit=1)
        if not banco_transf or not banco_transf.account_number:
            raise UserError(
                'No hay cuenta destino configurada para el servicio de Transferencia. '
                'Configure la cuenta en el servicio Transferencia / Crédito Inmediato '
                'de la pasarela antes de ejecutar las pruebas.'
            )

        self.state = 'running'
        self.result_line_ids.unlink()

        config = self.gateway_config_id
        client = config.get_client()

        tests = self._build_test_list(client, config, pos_session)
        ResultLine = self.env['ve.certification.result.line']

        for test_def in tests:
            line_vals = self._run_single_test(test_def, client, config, pos_session)
            line_vals['wizard_id'] = self.id
            ResultLine.create(line_vals)

        self.state = 'done'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _build_test_list(self, client, config, pos_session):
        """Retorna lista de dicts con la definición de cada prueba."""
        cid = self.ci_integrador
        telefono = self.telefono_integrador
        banco_p2c_code = self.banco_p2c_id.bank_id.code if self.banco_p2c_id else ''
        telefono_comercio_p2c = ''
        if self.banco_p2c_id and self.banco_p2c_id.phone_number:
            telefono_comercio_p2c = self.banco_p2c_id.phone_number

        # Datos de transferencia — obtenidos automáticamente desde Odoo
        banco_transf = self.env['ve.payment.service.bank'].search([
            ('service_id.gateway_config_id', '=', config.id),
            ('service_id.service_code', '=', 'transferencia'),
            ('active', '=', True),
        ], limit=1)
        cuenta_transf = banco_transf.account_number if banco_transf else ''

        # Datos de Zelle — banco del comercio desde la selección del wizard
        codigo_banco_zelle = self.banco_zelle_id.bank_id.code if self.banco_zelle_id else 'BOFA'

        return [
            # 1-3: Tarjeta
            {
                'test_id': 1, 'test_name': 'Tarjeta - Aprobada 0.01',
                'service_code': 'tarjeta', 'amount': '0.01',
                'method': 'compra_tarjeta',
                'params': dict(pan='5420070695259279', cvv2='000', expdate='1230',
                               amount='0.01', cid=cid, client='Test', factura='CERT-01'),
                'expect_approved': True,
            },
            {
                'test_id': 2, 'test_name': 'Tarjeta - Aprobada 10100.51',
                'service_code': 'tarjeta', 'amount': '10100.51',
                'method': 'compra_tarjeta',
                'params': dict(pan='5420070695259279', cvv2='000', expdate='1230',
                               amount='10100.51', cid=cid, client='Test', factura='CERT-02'),
                'expect_approved': True,
            },
            {
                'test_id': 3, 'test_name': 'Tarjeta - Rechazada 33500.01',
                'service_code': 'tarjeta', 'amount': '33500.01',
                'method': 'compra_tarjeta',
                'params': dict(pan='5420070695259279', cvv2='000', expdate='1230',
                               amount='33500.01', cid=cid, client='Test', factura='CERT-03'),
                'expect_approved': False,
            },
            # 4: PG Inactivo
            {
                'test_id': 4, 'test_name': 'PG Inactivo - Error controlado',
                'service_code': 'tarjeta', 'amount': '',
                'method': 'pg_inactivo',
                'params': {},
                'expect_approved': None,  # special case
            },
            # 5-7: P2C
            {
                'test_id': 5, 'test_name': 'P2C - Aprobada 1000',
                'service_code': 'p2c', 'amount': '1000.00',
                'method': 'pago_movil_p2c',
                'params': dict(telefonoCliente='04121234569', codigobancoCliente='0138',
                               codigobancoComercio=banco_p2c_code,
                               telefonoComercio=telefono_comercio_p2c,
                               amount='1000.00',
                               factura='CERT-05'),
                'expect_approved': True,
            },
            {
                'test_id': 6, 'test_name': 'P2C - Aprobada 25300.02',
                'service_code': 'p2c', 'amount': '25300.02',
                'method': 'pago_movil_p2c',
                'params': dict(telefonoCliente='04121234571', codigobancoCliente='0138',
                               codigobancoComercio=banco_p2c_code,
                               telefonoComercio=telefono_comercio_p2c,
                               amount='25300.02',
                               factura='CERT-06'),
                'expect_approved': True,
            },
            {
                'test_id': 7, 'test_name': 'P2C - Rechazada 25300.03',
                'service_code': 'p2c', 'amount': '25300.03',
                'method': 'pago_movil_p2c',
                'params': dict(telefonoCliente='04121234572', codigobancoCliente='0138',
                               codigobancoComercio=banco_p2c_code,
                               telefonoComercio=telefono_comercio_p2c,
                               amount='25300.03',
                               factura='CERT-07'),
                'expect_approved': False,
            },
            # 8-10: C2P
            {
                'test_id': 8, 'test_name': 'C2P - Aprobada 33500.01',
                'service_code': 'c2p', 'amount': '33500.01',
                'method': 'pago_movil_c2p',
                'params': dict(telefono=telefono, codigobanco='0138',
                               codigoc2p='12345678', amount='33500.01',
                               cid=cid, factura='CERT-08'),
                'expect_approved': True,
            },
            {
                'test_id': 9, 'test_name': 'C2P - Aprobada 10100.51',
                'service_code': 'c2p', 'amount': '10100.51',
                'method': 'pago_movil_c2p',
                'params': dict(telefono=telefono, codigobanco='0138',
                               codigoc2p='12345678', amount='10100.51',
                               cid=cid, factura='CERT-09'),
                'expect_approved': True,
            },
            {
                'test_id': 10, 'test_name': 'C2P - Rechazada 100000',
                'service_code': 'c2p', 'amount': '100000.00',
                'method': 'pago_movil_c2p',
                'params': dict(telefono=telefono, codigobanco='0138',
                               codigoc2p='12345678', amount='100000.00',
                               cid=cid, factura='CERT-10'),
                'expect_approved': False,
            },
            # 11-12: Vuelto
            {
                'test_id': 11, 'test_name': 'Vuelto - Aprobada 100',
                'service_code': 'vuelto', 'amount': '100.00',
                'method': 'vuelto_pago_movil',
                'params': dict(telefono=telefono, codigobanco='0138',
                               amount='100.00', cid=cid, factura='CERT-11'),
                'expect_approved': True,
            },
            {
                'test_id': 12, 'test_name': 'Vuelto - Aprobada 10100.51',
                'service_code': 'vuelto', 'amount': '10100.51',
                'method': 'vuelto_pago_movil',
                'params': dict(telefono=telefono, codigobanco='0138',
                               amount='10100.51', cid=cid, factura='CERT-12'),
                'expect_approved': True,
            },
            # 13-14: Transferencia (credito_inmediato)
            # Datos según corrección MegaSoft: cuentaOrigen puede ser 16 dígitos (PAN),
            # telefonoOrigen vacío, cid de prueba 'v6457425', banco origen '0105'
            {
                'test_id': 13, 'test_name': 'Transferencia - Aprobada 0.01',
                'service_code': 'transferencia', 'amount': '0.01',
                'method': 'credito_inmediato',
                'params': dict(cuentaOrigen='5420070695259279',
                               telefonoOrigen=telefono,
                               codigobancoOrigen='0105',
                               cuentaDestino=cuenta_transf,
                               amount='0.01', cid='V6457425', factura='CERT-13'),
                'expect_approved': True,
            },
            {
                'test_id': 14, 'test_name': 'Transferencia - Rechazada 33500.01',
                'service_code': 'transferencia', 'amount': '33500.01',
                'method': 'credito_inmediato',
                'params': dict(cuentaOrigen='5420070695259279',
                               telefonoOrigen=telefono,
                               codigobancoOrigen='0105',
                               cuentaDestino=cuenta_transf,
                               amount='33500.01', cid='V6457425', factura='CERT-14'),
                'expect_approved': False,
            },
            # 15-16: Zelle
            # Datos según corrección MegaSoft: codigobancoComercio debe ser 'BOFA' (no 'CHAS'),
            # cid de prueba 'V6457425'
            {
                'test_id': 15, 'test_name': 'Zelle - Aprobada 100000',
                'service_code': 'zelle', 'amount': '100000.00',
                'method': 'zelle',
                'params': dict(cid='V6457425',
                               codigobancoComercio=codigo_banco_zelle,
                               referencia='CERTZELLE1',
                               amount='100000.00',
                               factura='CERT-15'),
                'expect_approved': True,
            },
            {
                'test_id': 16, 'test_name': 'Zelle - Rechazada 33500.01',
                'service_code': 'zelle', 'amount': '33500.01',
                'method': 'zelle',
                'params': dict(cid='V6457425',
                               codigobancoComercio=codigo_banco_zelle,
                               referencia='CERTZELLE2',
                               amount='33500.01',
                               factura='CERT-16'),
                'expect_approved': False,
            },
            # 17-18: Crypto BTC (solicitud + confirmación automática)
            {
                'test_id': 17, 'test_name': 'Crypto BTC - Aprobada 4000000',
                'service_code': 'crypto', 'amount': '4000000.00',
                'method': 'crypto_solicitud',
                'follow_up': 'crypto_confirmacion',
                'params': dict(tipomoneda='BTC', amount='4000000.00',
                               factura='CERT-17'),
                'expect_approved': True,
            },
            {
                'test_id': 18, 'test_name': 'Crypto BTC - Rechazada 33500.01',
                'service_code': 'crypto', 'amount': '33500.01',
                'method': 'crypto_solicitud',
                'follow_up': 'crypto_confirmacion',
                'params': dict(tipomoneda='BTC', amount='33500.01',
                               factura='CERT-18'),
                'expect_approved': False,
            },
        ]

    def _run_single_test(self, test_def, client, config, pos_session):
        """Ejecuta una prueba individual y retorna vals para result.line."""
        test_id = test_def['test_id']
        line_vals = {
            'test_id': test_id,
            'test_name': test_def['test_name'],
            'service_code': test_def['service_code'],
            'amount_str': test_def.get('amount', ''),
            'passed': False,
            'error_detail': '',
            'requires_manual': test_def.get('requires_manual', False),
        }

        try:
            method_name = test_def['method']

            # Test 4: PG Inactivo
            if method_name == 'pg_inactivo':
                return self._test_pg_inactivo(line_vals)

            # Preregistro
            prereg = client.preregistro()
            prereg_request_xml = prereg.get('_request_xml', '')
            prereg_response_xml = prereg.get('_raw_xml', '')
            if prereg.get('error') or prereg.get('codigo') != '00':
                line_vals['error_detail'] = 'Preregistro falló: %s' % (
                    prereg.get('error') or prereg.get('descripcion', 'Error desconocido')
                )
                line_vals['request_xml'] = prereg_request_xml
                line_vals['response_xml'] = prereg_response_xml
                return line_vals

            control = prereg.get('control', '')
            line_vals['control'] = control

            # Ejecutar método
            params = {**test_def['params'], 'control': control}
            method = getattr(client, method_name)
            result = method(**params)

            # Capturar XML de solicitud y respuesta
            line_vals['request_xml'] = result.get('_request_xml', '')
            line_vals['response_xml'] = result.get('_raw_xml', '')
            line_vals['prereg_request_xml'] = prereg_request_xml
            line_vals['prereg_response_xml'] = prereg_response_xml

            # Registrar resultado
            line_vals['referencia'] = result.get('referencia', '')
            line_vals['codigo'] = result.get('codigo', '')
            line_vals['descripcion'] = result.get('descripcion', '')
            line_vals['voucher'] = result.get('voucher', '')

            # Crypto: solicitud + confirmación automática
            if test_def.get('follow_up') == 'crypto_confirmacion':
                qr_url = result.get('qrurl', '')
                line_vals['qr_url'] = qr_url
                # Paso 2: Confirmación automática con el mismo control
                try:
                    confirm_result = client.crypto_confirmacion(control=control)
                    line_vals['response_xml'] = confirm_result.get('_raw_xml', '')
                    result = confirm_result  # Usar resultado de confirmación
                    line_vals['referencia'] = confirm_result.get('referencia', '') or line_vals.get('referencia', '')
                    line_vals['codigo'] = confirm_result.get('codigo', '')
                    line_vals['descripcion'] = confirm_result.get('descripcion', '')
                    line_vals['voucher'] = confirm_result.get('voucher', '')
                except Exception as e:
                    line_vals['error_detail'] = 'Error en crypto_confirmacion: %s' % str(e)
                    return line_vals

            # Evaluar PASS/FAIL
            codigo = line_vals.get('codigo') or result.get('codigo', '')
            if test_def['expect_approved']:
                line_vals['passed'] = (codigo == '00')
                if not line_vals['passed']:
                    line_vals['error_detail'] = 'Esperaba código 00, recibió: %s' % codigo
            else:
                line_vals['passed'] = (codigo != '00')
                if not line_vals['passed']:
                    line_vals['error_detail'] = 'Esperaba rechazo, pero recibió código 00'

            # Registrar en log
            try:
                log = self.env['ve.bank.transaction.log'].sudo().create_from_gateway_response(
                    vals={**result, 'amount': test_def.get('amount', ''),
                          'factura': params.get('factura', '')},
                    gateway_config=config,
                    service_code=test_def['service_code'],
                    pos_session=pos_session if test_def['service_code'] != 'tarjeta' else None,
                )
                line_vals['log_id'] = log.id if log else False
            except Exception as e:
                _logger.warning('Error creando log para test %s: %s', test_id, e)

        except Exception as e:
            line_vals['error_detail'] = str(e)
            _logger.warning('Error en prueba %s: %s', test_id, e, exc_info=True)

        return line_vals

    def _test_pg_inactivo(self, line_vals):
        """Test 4: PG con URL inválida."""
        try:
            fake_config = PGConfig(
                base_url='https://pg-inactivo-test.invalid',
                usuario='test',
                contrasena='test',
                codafiliacion='00000000',
            )
            fake_client = PaymentGatewayClient(fake_config, timeout=5)
            result = fake_client.preregistro()

            if result.get('error'):
                line_vals['passed'] = True
                line_vals['error_detail'] = 'Error capturado: %s' % result.get('error', '')
                line_vals['codigo'] = result.get('codigo', '')
            else:
                line_vals['passed'] = False
                line_vals['error_detail'] = 'Se esperaba error pero se recibió respuesta exitosa'
        except Exception as e:
            line_vals['passed'] = True
            line_vals['error_detail'] = 'Excepción capturada: %s' % str(e)

        return line_vals

    # ── Exportar a Excel ──────────────────────────────────────────

    def action_export_excel(self):
        self.ensure_one()
        if not Workbook:
            raise UserError('Se requiere la librería openpyxl para exportar a Excel.')

        wb = Workbook()

        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = 'Resumen'
        header_font = Font(bold=True, size=12)
        pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        ws1.append(['Certificación MegaSoft - Resultados'])
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(['Configuración:', self.gateway_config_id.name or ''])
        ws1.append(['CI Integrador:', self.ci_integrador])
        ws1.append(['Fecha:', fields.Datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws1.append(['Total:', self.total_count, 'PASS:', self.passed_count, 'FAIL:', self.failed_count])
        ws1.append([])

        # Tabla de resultados
        headers = ['#', 'Prueba', 'Servicio', 'Monto', 'Código', 'Referencia', 'Resultado', 'Detalle']
        ws1.append(headers)
        for cell in ws1[ws1.max_row]:
            cell.font = header_font

        for line in self.result_line_ids:
            row = [
                line.test_id, line.test_name, line.service_code,
                line.amount_str, line.codigo or '', line.referencia or '',
                'PASS' if line.passed else 'FAIL', line.error_detail or '',
            ]
            ws1.append(row)
            result_cell = ws1.cell(row=ws1.max_row, column=7)
            result_cell.fill = pass_fill if line.passed else fail_fill

        for col in ws1.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        # Hoja 2: Vouchers
        ws2 = wb.create_sheet('Vouchers')
        ws2.append(['#', 'Prueba', 'Voucher'])
        for cell in ws2[1]:
            cell.font = header_font
        mono_font = Font(name='Courier New', size=9)
        for line in self.result_line_ids.filtered(lambda l: l.voucher):
            ws2.append([line.test_id, line.test_name, line.voucher])
            ws2.cell(row=ws2.max_row, column=3).font = mono_font
            ws2.cell(row=ws2.max_row, column=3).alignment = Alignment(wrap_text=True)

        # Hoja 3: Datos Tecnicos
        ws3 = wb.create_sheet('Datos Técnicos')
        ws3.append(['#', 'Prueba', 'Servicio', 'Control', 'Código', 'Descripción', 'Referencia'])
        for cell in ws3[1]:
            cell.font = header_font
        for line in self.result_line_ids:
            ws3.append([
                line.test_id, line.test_name, line.service_code,
                line.control or '', line.codigo or '',
                line.descripcion or '', line.referencia or '',
            ])

        # Hoja 4: Solicitudes y Respuestas
        ws4 = wb.create_sheet('Solicitudes y Respuestas')
        detail_headers = [
            '#', 'Prueba', 'Servicio', 'Resultado',
            'PreRegistro - Solicitud XML', 'PreRegistro - Respuesta XML',
            'Solicitud XML', 'Respuesta XML',
            'Código', 'Descripción', 'Control', 'Referencia', 'Detalle Error',
        ]
        ws4.append(detail_headers)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, size=10, color='FFFFFF')
        for cell in ws4[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for line in self.result_line_ids:
            row = [
                line.test_id,
                line.test_name,
                line.service_code,
                'PASS' if line.passed else 'FAIL',
                line.prereg_request_xml or '',
                line.prereg_response_xml or '',
                line.request_xml or '',
                line.response_xml or '',
                line.codigo or '',
                line.descripcion or '',
                line.control or '',
                line.referencia or '',
                line.error_detail or '',
            ]
            ws4.append(row)
            # Colorear resultado
            result_cell = ws4.cell(row=ws4.max_row, column=4)
            result_cell.fill = pass_fill if line.passed else fail_fill
            # Formato monoespaciado para XML
            for col_idx in (5, 6, 7, 8):
                xml_cell = ws4.cell(row=ws4.max_row, column=col_idx)
                xml_cell.font = mono_font
                xml_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Ajustar anchos hoja 4
        ws4.column_dimensions['A'].width = 5
        ws4.column_dimensions['B'].width = 30
        ws4.column_dimensions['C'].width = 14
        ws4.column_dimensions['D'].width = 10
        ws4.column_dimensions['E'].width = 50
        ws4.column_dimensions['F'].width = 50
        ws4.column_dimensions['G'].width = 60
        ws4.column_dimensions['H'].width = 60
        ws4.column_dimensions['I'].width = 10
        ws4.column_dimensions['J'].width = 25
        ws4.column_dimensions['K'].width = 22
        ws4.column_dimensions['L'].width = 15
        ws4.column_dimensions['M'].width = 35

        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        self.export_file = base64.b64encode(output.read())
        self.export_filename = 'certificacion_megasoft_%s.xlsx' % (
            fields.Date.today().strftime('%Y%m%d')
        )

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }


class VeCertificationResultLine(models.TransientModel):
    _name = 've.certification.result.line'
    _description = 'Resultado de Prueba de Certificación'
    _order = 'test_id'

    wizard_id = fields.Many2one('ve.certification.wizard', ondelete='cascade')
    test_id = fields.Integer(string='#')
    test_name = fields.Char(string='Prueba')
    service_code = fields.Char(string='Servicio')
    amount_str = fields.Char(string='Monto')
    control = fields.Char(string='Control')
    referencia = fields.Char(string='Referencia')
    codigo = fields.Char(string='Código')
    descripcion = fields.Char(string='Descripción')
    voucher = fields.Text(string='Voucher')
    passed = fields.Boolean(string='PASS')
    error_detail = fields.Text(string='Detalle')
    log_id = fields.Many2one('ve.bank.transaction.log', string='Log Generado')
    requires_manual = fields.Boolean(string='Requiere acción manual')
    qr_url = fields.Char(string='URL QR (Crypto)')
    request_xml = fields.Text(string='Solicitud XML')
    response_xml = fields.Text(string='Respuesta XML')
    prereg_request_xml = fields.Text(string='PreRegistro Solicitud XML')
    prereg_response_xml = fields.Text(string='PreRegistro Respuesta XML')
